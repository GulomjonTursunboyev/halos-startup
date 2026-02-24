"""
HALOS PRO Features Module
Statistics, Reminders, Debt Monitoring, Excel Export
"""
import logging
from datetime import datetime, timedelta
from typing import Dict, Any, List, Optional
from telegram import (
    Update,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
)
from telegram.ext import ContextTypes
from telegram.constants import ParseMode

from app.database import get_database
from app.languages import get_message, format_number
from app.subscription_handlers import is_user_pro

logger = logging.getLogger(__name__)


async def get_pro_trial_buttons(telegram_id: int, lang: str) -> List[List[InlineKeyboardButton]]:
    """Generate PRO and Trial buttons based on user status"""
    db = await get_database()
    user = await db.get_user(telegram_id)
    
    buttons = []
    
    # Check if trial is available
    trial_available = user and not user.get("trial_used", 0)
    
    if trial_available:
        # Show both trial (highlighted) and PRO buttons
        buttons.append([InlineKeyboardButton(
            "🎁 3 kun BEPUL sinash" if lang == "uz" else "🎁 3 дня БЕСПЛАТНО",
            callback_data="activate_trial"
        )])
    
    buttons.append([InlineKeyboardButton(
        "💎 PRO olish" if lang == "uz" else "💎 Получить PRO",
        callback_data="show_pricing"
    )])
    
    return buttons


# ==================== STATISTICS ====================

async def show_statistics(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Show user statistics - weekly/monthly/yearly"""
    query = update.callback_query
    if query:
        await query.answer()
    
    telegram_id = update.effective_user.id
    lang = context.user_data.get("lang", "uz")
    
    # Check PRO status
    is_pro = await is_user_pro(telegram_id)
    if not is_pro:
        if lang == "uz":
            msg = "🔒 *Statistika PRO foydalanuvchilar uchun*\n\nPRO ga o'ting va batafsil statistikani ko'ring!"
        else:
            msg = "🔒 *Статистика для PRO пользователей*\n\nПерейдите на PRO и смотрите детальную статистику!"
        
        keyboard = await get_pro_trial_buttons(telegram_id, lang)
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        if query:
            await query.edit_message_text(msg, parse_mode="Markdown", reply_markup=reply_markup)
        else:
            await update.message.reply_text(msg, parse_mode="Markdown", reply_markup=reply_markup)
        return
    
    # Get user data
    db = await get_database()
    user = await db.get_user(telegram_id)
    profile = await db.get_financial_profile(user["id"]) if user else None
    
    if not profile:
        if lang == "uz":
            msg = "📊 Statistika ko'rish uchun avval ma'lumotlaringizni kiriting."
        else:
            msg = "📊 Для просмотра статистики сначала введите свои данные."
        
        if query:
            await query.edit_message_text(msg)
        else:
            await update.message.reply_text(msg)
        return
    
    # Build statistics
    income = profile.get("income_self", 0) + profile.get("income_partner", 0)
    mandatory = profile.get("rent", 0) + profile.get("kindergarten", 0) + profile.get("utilities", 0)
    loan_payment = profile.get("loan_payment", 0)
    total_debt = profile.get("total_debt", 0)
    free_cash = income - mandatory - loan_payment
    
    if free_cash > 0:
        savings = free_cash * 0.1  # 10% savings
        extra_debt = free_cash * 0.2  # 20% extra to debt
        living = free_cash * 0.7  # 70% for living
    else:
        savings = living = extra_debt = 0
    
    # Calculate projections
    weekly_savings = savings / 4
    monthly_savings = savings
    yearly_savings = savings * 12
    
    weekly_income = income / 4
    monthly_income = income
    yearly_income = income * 12
    
    weekly_expense = (mandatory + loan_payment + living) / 4
    monthly_expense = mandatory + loan_payment + living
    yearly_expense = monthly_expense * 12
    
    # Get AI transaction summary
    from app.ai_assistant import get_transaction_summary
    ai_summary = await get_transaction_summary(db, user["id"], days=30)
    ai_income = ai_summary.get("total_income", 0)
    ai_expense = ai_summary.get("total_expense", 0)
    ai_balance = ai_income - ai_expense
    
    if lang == "uz":
        msg = (
            "📊 *SIZNING STATISTIKANGIZ*\n"
            "━━━━━━━━━━━━━━━━━━━━\n\n"
            
            "📅 *HAFTALIK:*\n"
            f"├ 💰 Daromad: ~{format_number(int(weekly_income))} so'm\n"
            f"├ 💸 Xarajat: ~{format_number(int(weekly_expense))} so'm\n"
            f"└ 🏦 Boylik: +{format_number(int(weekly_savings))} so'm\n\n"
            
            "📅 *OYLIK:*\n"
            f"├ 💰 Daromad: {format_number(int(monthly_income))} so'm\n"
            f"├ 💸 Xarajat: {format_number(int(monthly_expense))} so'm\n"
            f"└ 🏦 Boylik: +{format_number(int(monthly_savings))} so'm\n\n"
            
            "📅 *YILLIK PROGNOZ:*\n"
            f"├ 💰 Daromad: {format_number(int(yearly_income))} so'm\n"
            f"├ 💸 Xarajat: {format_number(int(yearly_expense))} so'm\n"
            f"└ 🏦 Boylik: +{format_number(int(yearly_savings))} so'm\n\n"
            
            "━━━━━━━━━━━━━━━━━━━━\n"
            "🤖 *AI YORDAMCHI (30 kun):*\n"
            "━━━━━━━━━━━━━━━━━━━━\n"
            f"├ 💰 Daromadlar: +{format_number(int(ai_income))} so'm\n"
            f"├ 💸 Xarajatlar: -{format_number(int(ai_expense))} so'm\n"
            f"└ 📊 Balans: {'+' if ai_balance >= 0 else ''}{format_number(int(ai_balance))} so'm\n\n"
            
            "━━━━━━━━━━━━━━━━━━━━\n"
            f"💳 Umumiy qarz: {format_number(int(total_debt))} so'm"
        )
    else:
        msg = (
            "📊 *ВАША СТАТИСТИКА*\n"
            "━━━━━━━━━━━━━━━━━━━━\n\n"
            
            "📅 *ЕЖЕНЕДЕЛЬНО:*\n"
            f"├ 💰 Доход: ~{format_number(int(weekly_income))} сум\n"
            f"├ 💸 Расход: ~{format_number(int(weekly_expense))} сум\n"
            f"└ 🏦 Богатство: +{format_number(int(weekly_savings))} сум\n\n"
            
            "📅 *ЕЖЕМЕСЯЧНО:*\n"
            f"├ 💰 Доход: {format_number(int(monthly_income))} сум\n"
            f"├ 💸 Расход: {format_number(int(monthly_expense))} сум\n"
            f"└ 🏦 Богатство: +{format_number(int(monthly_savings))} сум\n\n"
            
            "📅 *ГОДОВОЙ ПРОГНОЗ:*\n"
            f"├ 💰 Доход: {format_number(int(yearly_income))} сум\n"
            f"├ 💸 Расход: {format_number(int(yearly_expense))} сум\n"
            f"└ 🏦 Богатство: +{format_number(int(yearly_savings))} сум\n\n"
            
            "━━━━━━━━━━━━━━━━━━━━\n"
            "🤖 *AI ПОМОЩНИК (30 дней):*\n"
            "━━━━━━━━━━━━━━━━━━━━\n"
            f"├ 💰 Доходы: +{format_number(int(ai_income))} сум\n"
            f"├ 💸 Расходы: -{format_number(int(ai_expense))} сум\n"
            f"└ 📊 Баланс: {'+' if ai_balance >= 0 else ''}{format_number(int(ai_balance))} сум\n\n"
            
            "━━━━━━━━━━━━━━━━━━━━\n"
            f"💳 Общий долг: {format_number(int(total_debt))} сум"
        )
    
    keyboard = [[InlineKeyboardButton(
        "◀️ Orqaga" if lang == "uz" else "◀️ Назад",
        callback_data="pro_menu"
    )]]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    if query:
        await query.edit_message_text(msg, parse_mode="Markdown", reply_markup=reply_markup)
    else:
        await update.message.reply_text(msg, parse_mode="Markdown", reply_markup=reply_markup)


# ==================== REMINDERS ====================

async def show_reminders(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Show and manage payment reminders"""
    query = update.callback_query
    if query:
        await query.answer()
    
    telegram_id = update.effective_user.id
    lang = context.user_data.get("lang", "uz")
    
    # Check PRO status
    is_pro = await is_user_pro(telegram_id)
    if not is_pro:
        if lang == "uz":
            msg = "🔒 *Eslatmalar PRO foydalanuvchilar uchun*\n\nPRO ga o'ting va to'lov eslatmalarini oling!"
        else:
            msg = "🔒 *Напоминания для PRO пользователей*\n\nПерейдите на PRO и получайте напоминания!"
        
        keyboard = await get_pro_trial_buttons(telegram_id, lang)
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        if query:
            await query.edit_message_text(msg, parse_mode="Markdown", reply_markup=reply_markup)
        else:
            await update.message.reply_text(msg, parse_mode="Markdown", reply_markup=reply_markup)
        return
    
    # Get user's reminders
    db = await get_database()
    user = await db.get_user(telegram_id)
    
    if lang == "uz":
        msg = (
            "🔔 *TO'LOV ESLATMALARI*\n"
            "━━━━━━━━━━━━━━━━━━━━\n\n"
            
            "HALOS sizga quyidagilarni eslatib turadi:\n\n"
            
            "✅ *Oylik to'lov* — har oy 1-sanada\n"
            "✅ *Haftalik hisobot* — har dushanba\n"
            "✅ *Qarz holati* — har 15 kunda\n"
            "✅ *Boylik o'sishi* — har oyda\n\n"
            
            "━━━━━━━━━━━━━━━━━━━━\n"
            "💡 Eslatmalar avtomatik yoqilgan.\n"
            "Xabarnomalar Telegram orqali keladi."
        )
    else:
        msg = (
            "🔔 *НАПОМИНАНИЯ ОБ ОПЛАТЕ*\n"
            "━━━━━━━━━━━━━━━━━━━━\n\n"
            
            "HALOS напомнит вам о:\n\n"
            
            "✅ *Ежемесячный платёж* — 1-го числа\n"
            "✅ *Еженедельный отчёт* — каждый понедельник\n"
            "✅ *Состояние долга* — каждые 15 дней\n"
            "✅ *Рост богатства* — каждый месяц\n\n"
            
            "━━━━━━━━━━━━━━━━━━━━\n"
            "💡 Напоминания включены автоматически.\n"
            "Уведомления приходят через Telegram."
        )
    
    keyboard = [
        [InlineKeyboardButton(
            "🔕 O'chirish" if lang == "uz" else "🔕 Отключить",
            callback_data="toggle_reminders_off"
        ), InlineKeyboardButton(
            "🔔 Yoqish" if lang == "uz" else "🔔 Включить",
            callback_data="toggle_reminders_on"
        )],
        [InlineKeyboardButton(
            "◀️ Orqaga" if lang == "uz" else "◀️ Назад",
            callback_data="pro_menu"
        )]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    if query:
        await query.edit_message_text(msg, parse_mode="Markdown", reply_markup=reply_markup)
    else:
        await update.message.reply_text(msg, parse_mode="Markdown", reply_markup=reply_markup)


# ==================== DEBT MONITORING ====================

async def show_debt_monitoring(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Show debt monitoring with progress and tips"""
    query = update.callback_query
    if query:
        await query.answer()
    
    telegram_id = update.effective_user.id
    lang = context.user_data.get("lang", "uz")
    
    # Check PRO status
    is_pro = await is_user_pro(telegram_id)
    if not is_pro:
        if lang == "uz":
            msg = "🔒 *Qarz nazorati PRO foydalanuvchilar uchun*\n\nPRO ga o'ting va qarzlaringizni nazorat qiling!"
        else:
            msg = "🔒 *Контроль долгов для PRO пользователей*\n\nПерейдите на PRO и контролируйте долги!"
        
        keyboard = await get_pro_trial_buttons(telegram_id, lang)
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        if query:
            await query.edit_message_text(msg, parse_mode="Markdown", reply_markup=reply_markup)
        else:
            await update.message.reply_text(msg, parse_mode="Markdown", reply_markup=reply_markup)
        return
    
    # Get user data
    db = await get_database()
    user = await db.get_user(telegram_id)
    profile = await db.get_financial_profile(user["id"]) if user else None
    
    if not profile:
        if lang == "uz":
            msg = "📋 Qarz nazorati uchun avval ma'lumotlaringizni kiriting."
        else:
            msg = "📋 Для контроля долгов сначала введите свои данные."
        
        if query:
            await query.edit_message_text(msg)
        else:
            await update.message.reply_text(msg)
        return
    
    # Calculate debt progress
    total_debt = profile.get("total_debt", 0)
    loan_payment = profile.get("loan_payment", 0)
    income = profile.get("income_self", 0) + profile.get("income_partner", 0)
    mandatory = profile.get("rent", 0) + profile.get("kindergarten", 0) + profile.get("utilities", 0)
    free_cash = income - mandatory - loan_payment
    
    import math
    if loan_payment > 0 and total_debt > 0:
        # Simple months
        simple_months = math.ceil(total_debt / loan_payment)
        
        # PRO months (with extra payments)
        if free_cash > 0:
            extra_debt = free_cash * 0.2
            total_payment = loan_payment + extra_debt
            pro_months = math.ceil(total_debt / total_payment)
        else:
            pro_months = simple_months
            extra_debt = 0
        
        # Calculate progress (assume user started today)
        progress_percent = 0  # Will be calculated based on actual tracking
        
        # Exit dates
        today = datetime.now()
        simple_exit = today + timedelta(days=simple_months * 30)
        pro_exit = today + timedelta(days=pro_months * 30)
        
        # Progress bar
        bar_filled = int(progress_percent / 10)
        bar_empty = 10 - bar_filled
        progress_bar = "█" * bar_filled + "░" * bar_empty
        
        if lang == "uz":
            msg = (
                "📋 *YO'LINGIZ NAZORATI*\n"
                "━━━━━━━━━━━━━━━━━━━━\n\n"
                
                f"💳 *Umumiy yuk:* {format_number(int(total_debt))} so'm\n"
                f"📅 *Oylik to'lov:* {format_number(int(loan_payment))} so'm\n"
                f"⚡ *Qo'shimcha:* +{format_number(int(extra_debt))} so'm\n\n"
                
                f"[{progress_bar}] {progress_percent}%\n\n"
                
                "━━━━━━━━━━━━━━━━━━━━\n"
                "📊 *PROGNOZ:*\n"
                f"├ Oddiy yo'l: {simple_exit.strftime('%B %Y')}\n"
                f"└ HALOS PRO bilan: {pro_exit.strftime('%B %Y')}\n\n"
                
                "━━━━━━━━━━━━━━━━━━━━\n"
                "💡 *MASLAHAT:*\n"
                "Har oyda qo'shimcha to'lov qilsangiz,\n"
                f"*{simple_months - pro_months} oy* oldin yengillikka erishasiz!"
            )
        else:
            msg = (
                "📋 *КОНТРОЛЬ ВАШЕГО ПУТИ*\n"
                "━━━━━━━━━━━━━━━━━━━━\n\n"
                
                f"💳 *Общее бремя:* {format_number(int(total_debt))} сум\n"
                f"📅 *Ежемесячно:* {format_number(int(loan_payment))} сум\n"
                f"⚡ *Дополнительно:* +{format_number(int(extra_debt))} сум\n\n"
                
                f"[{progress_bar}] {progress_percent}%\n\n"
                
                "━━━━━━━━━━━━━━━━━━━━\n"
                "📊 *ПРОГНОЗ:*\n"
                f"├ Обычный путь: {simple_exit.strftime('%B %Y')}\n"
                f"└ С HALOS PRO: {pro_exit.strftime('%B %Y')}\n\n"
                
                "━━━━━━━━━━━━━━━━━━━━\n"
                "💡 *СОВЕТ:*\n"
                "Делая дополнительные платежи,\n"
                f"достигнете свободы на *{simple_months - pro_months} мес* раньше!"
            )
    else:
        if lang == "uz":
            msg = "📋 Qarz ma'lumotlari topilmadi. Iltimos, qarz summasini kiriting."
        else:
            msg = "📋 Данные о долге не найдены. Пожалуйста, введите сумму долга."
    
    keyboard = [[InlineKeyboardButton(
        "◀️ Orqaga" if lang == "uz" else "◀️ Назад",
        callback_data="pro_menu"
    )]]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    if query:
        await query.edit_message_text(msg, parse_mode="Markdown", reply_markup=reply_markup)
    else:
        await update.message.reply_text(msg, parse_mode="Markdown", reply_markup=reply_markup)


# ==================== EXCEL EXPORT ====================

async def export_excel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Export user data to beautifully formatted Excel file"""
    query = update.callback_query
    if query:
        await query.answer()
    
    telegram_id = update.effective_user.id
    lang = context.user_data.get("lang", "uz")
    
    # Check PRO status
    is_pro = await is_user_pro(telegram_id)
    if not is_pro:
        if lang == "uz":
            msg = "🔒 *Excel eksport PRO foydalanuvchilar uchun*\n\nPRO ga o'ting va hisobotlaringizni yuklab oling!"
        else:
            msg = "🔒 *Excel экспорт для PRO пользователей*\n\nПерейдите на PRO и скачивайте отчёты!"
        
        keyboard = await get_pro_trial_buttons(telegram_id, lang)
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        if query:
            await query.edit_message_text(msg, parse_mode="Markdown", reply_markup=reply_markup)
        else:
            await update.message.reply_text(msg, parse_mode="Markdown", reply_markup=reply_markup)
        return
    
    # Show loading message immediately (before long Excel generation)
    loading_msg = "⏳ *Excel hisobot tayyorlanmoqda...*\n\n_Bu bir necha soniya vaqt olishi mumkin_" if lang == "uz" else "⏳ *Формируем Excel отчёт...*\n\n_Это может занять несколько секунд_"
    if query:
        await query.edit_message_text(loading_msg, parse_mode="Markdown")
    
    # Get user data
    db = await get_database()
    user = await db.get_user(telegram_id)
    profile = await db.get_financial_profile(user["id"]) if user else None
    
    if not profile:
        if lang == "uz":
            msg = "📥 Eksport qilish uchun avval ma'lumotlaringizni kiriting."
        else:
            msg = "📥 Для экспорта сначала введите свои данные."
        
        if query:
            await query.edit_message_text(msg)
        else:
            await update.message.reply_text(msg)
        return

    
    # Generate beautiful Excel file
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle
        from openpyxl.utils import get_column_letter
        from openpyxl.chart import PieChart, Reference, BarChart
        from openpyxl.chart.label import DataLabelList
        from io import BytesIO
        import math
        
        # Get transactions for the month
        from app.ai_assistant import get_transaction_summary
        month_summary = await get_transaction_summary(db, user["id"], days=30)
        
        # Prepare data
        income_self = profile.get("income_self", 0) or 0
        income_partner = profile.get("income_partner", 0) or 0
        total_income = income_self + income_partner
        
        rent = profile.get("rent", 0) or 0
        kindergarten = profile.get("kindergarten", 0) or 0
        utilities = profile.get("utilities", 0) or 0
        loan_payment = profile.get("loan_payment", 0) or 0
        total_debt = profile.get("total_debt", 0) or 0
        
        mandatory = rent + kindergarten + utilities + loan_payment
        free_cash = total_income - mandatory
        
        if free_cash > 0:
            savings = free_cash * 0.10
            extra_debt = free_cash * 0.20
            living = free_cash * 0.70
        else:
            savings = living = extra_debt = 0
        
        # Calculate exit months
        if loan_payment > 0 and total_debt > 0:
            simple_months = math.ceil(total_debt / loan_payment)
            total_payment = loan_payment + extra_debt
            pro_months = math.ceil(total_debt / total_payment) if total_payment > 0 else simple_months
            months_saved = simple_months - pro_months
        else:
            simple_months = pro_months = months_saved = 0
        
        # Transaction summary
        monthly_expense = month_summary.get("total_expense", 0)
        monthly_income_tx = month_summary.get("total_income", 0)
        expense_by_cat = month_summary.get("expense_by_category", {})
        
        # Create workbook
        wb = Workbook()
        
        # ==================== STYLES ====================
        # Header style
        header_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')  # Green
        header_align = Alignment(horizontal='center', vertical='center')
        
        # Title style
        title_font = Font(name='Arial', size=18, bold=True, color='1B5E20')
        title_align = Alignment(horizontal='center', vertical='center')
        
        # Section header
        section_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        section_fill = PatternFill(start_color='1976D2', end_color='1976D2', fill_type='solid')  # Blue
        
        # Data styles
        label_font = Font(name='Arial', size=11)
        value_font = Font(name='Arial', size=11, bold=True)
        money_font = Font(name='Arial', size=11, bold=True, color='2E7D32')
        danger_font = Font(name='Arial', size=11, bold=True, color='D32F2F')
        
        # Border
        thin_border = Border(
            left=Side(style='thin', color='BDBDBD'),
            right=Side(style='thin', color='BDBDBD'),
            top=Side(style='thin', color='BDBDBD'),
            bottom=Side(style='thin', color='BDBDBD')
        )
        
        # Fills
        income_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')  # Light green
        expense_fill = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')  # Light red
        highlight_fill = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')  # Light orange
        
        # ==================== SHEET 1: SUMMARY ====================
        ws1 = wb.active
        ws1.title = "Xulosa" if lang == "uz" else "Сводка"
        
        # Set column widths
        ws1.column_dimensions['A'].width = 35
        ws1.column_dimensions['B'].width = 20
        ws1.column_dimensions['C'].width = 15
        
        # Title
        ws1.merge_cells('A1:C1')
        ws1['A1'] = "📊 HALOS MOLIYAVIY HISOBOT" if lang == "uz" else "📊 HALOS ФИНАНСОВЫЙ ОТЧЁТ"
        ws1['A1'].font = title_font
        ws1['A1'].alignment = title_align
        ws1.row_dimensions[1].height = 30
        
        # Date
        ws1.merge_cells('A2:C2')
        ws1['A2'] = f"📅 {datetime.now().strftime('%d.%m.%Y')}"
        ws1['A2'].font = Font(name='Arial', size=10, italic=True)
        ws1['A2'].alignment = Alignment(horizontal='center')
        
        row = 4
        
        # ===== DAROMADLAR =====
        ws1.merge_cells(f'A{row}:C{row}')
        ws1[f'A{row}'] = "💰 DAROMADLAR" if lang == "uz" else "💰 ДОХОДЫ"
        ws1[f'A{row}'].font = section_font
        ws1[f'A{row}'].fill = section_fill
        ws1[f'A{row}'].alignment = header_align
        ws1.row_dimensions[row].height = 25
        row += 1
        
        income_data = [
            ("Shaxsiy daromad" if lang == "uz" else "Личный доход", income_self),
            ("Sherik daromadi" if lang == "uz" else "Доход партнёра", income_partner),
            ("JAMI DAROMAD" if lang == "uz" else "ОБЩИЙ ДОХОД", total_income),
        ]
        
        for label, value in income_data:
            ws1[f'A{row}'] = label
            ws1[f'A{row}'].font = label_font
            ws1[f'A{row}'].border = thin_border
            ws1[f'A{row}'].fill = income_fill
            
            ws1[f'B{row}'] = value
            ws1[f'B{row}'].font = money_font if "JAMI" in label or "ОБЩИЙ" in label else value_font
            ws1[f'B{row}'].number_format = '#,##0'
            ws1[f'B{row}'].border = thin_border
            ws1[f'B{row}'].fill = income_fill
            ws1[f'B{row}'].alignment = Alignment(horizontal='right')
            
            ws1[f'C{row}'] = "so'm" if lang == "uz" else "сум"
            ws1[f'C{row}'].border = thin_border
            ws1[f'C{row}'].fill = income_fill
            row += 1
        
        row += 1
        
        # ===== MAJBURIY XARAJATLAR =====
        ws1.merge_cells(f'A{row}:C{row}')
        ws1[f'A{row}'] = "🏠 MAJBURIY XARAJATLAR" if lang == "uz" else "🏠 ОБЯЗАТЕЛЬНЫЕ РАСХОДЫ"
        ws1[f'A{row}'].font = section_font
        ws1[f'A{row}'].fill = PatternFill(start_color='E65100', end_color='E65100', fill_type='solid')
        ws1[f'A{row}'].alignment = header_align
        ws1.row_dimensions[row].height = 25
        row += 1
        
        expense_data = [
            ("Ijara" if lang == "uz" else "Аренда", rent),
            ("Bog'cha" if lang == "uz" else "Детсад", kindergarten),
            ("Kommunal" if lang == "uz" else "Коммуналка", utilities),
            ("Kredit to'lovi" if lang == "uz" else "Платёж по кредиту", loan_payment),
            ("JAMI MAJBURIY" if lang == "uz" else "ИТОГО ОБЯЗАТ.", mandatory),
        ]
        
        for label, value in expense_data:
            ws1[f'A{row}'] = label
            ws1[f'A{row}'].font = label_font
            ws1[f'A{row}'].border = thin_border
            ws1[f'A{row}'].fill = expense_fill
            
            ws1[f'B{row}'] = value
            ws1[f'B{row}'].font = danger_font if "JAMI" in label or "ИТОГО" in label else value_font
            ws1[f'B{row}'].number_format = '#,##0'
            ws1[f'B{row}'].border = thin_border
            ws1[f'B{row}'].fill = expense_fill
            ws1[f'B{row}'].alignment = Alignment(horizontal='right')
            
            ws1[f'C{row}'] = "so'm" if lang == "uz" else "сум"
            ws1[f'C{row}'].border = thin_border
            ws1[f'C{row}'].fill = expense_fill
            row += 1
        
        row += 1
        
        # ===== BO'SH PUL TAQSIMOTI =====
        ws1.merge_cells(f'A{row}:C{row}')
        ws1[f'A{row}'] = "🌟 HALOS USULI" if lang == "uz" else "🌟 МЕТОД HALOS"
        ws1[f'A{row}'].font = section_font
        ws1[f'A{row}'].fill = PatternFill(start_color='7B1FA2', end_color='7B1FA2', fill_type='solid')
        ws1[f'A{row}'].alignment = header_align
        ws1.row_dimensions[row].height = 25
        row += 1
        
        budget_data = [
            ("Bo'sh pul" if lang == "uz" else "Свободные средства", free_cash),
            ("🏠 Yashash" if lang == "uz" else "🏠 Жизнь", living),
            ("⚡ Qarz to'lash" if lang == "uz" else "⚡ Погашение", extra_debt),
            ("💰 Jamg'arma" if lang == "uz" else "💰 Накопления", savings),
        ]
        
        for label, value in budget_data:
            ws1[f'A{row}'] = label
            ws1[f'A{row}'].font = label_font
            ws1[f'A{row}'].border = thin_border
            ws1[f'A{row}'].fill = highlight_fill
            
            ws1[f'B{row}'] = value
            ws1[f'B{row}'].font = money_font
            ws1[f'B{row}'].number_format = '#,##0'
            ws1[f'B{row}'].border = thin_border
            ws1[f'B{row}'].fill = highlight_fill
            ws1[f'B{row}'].alignment = Alignment(horizontal='right')
            
            ws1[f'C{row}'] = "so'm" if lang == "uz" else "сум"
            ws1[f'C{row}'].border = thin_border
            ws1[f'C{row}'].fill = highlight_fill
            ws1[f'C{row}'].alignment = Alignment(horizontal='center')
            row += 1
        
        row += 1
        
        # ===== QARZ MA'LUMOTLARI =====
        if total_debt > 0:
            ws1.merge_cells(f'A{row}:C{row}')
            ws1[f'A{row}'] = "📅 QARZDAN CHIQISH" if lang == "uz" else "📅 ВЫХОД ИЗ ДОЛГА"
            ws1[f'A{row}'].font = section_font
            ws1[f'A{row}'].fill = PatternFill(start_color='C62828', end_color='C62828', fill_type='solid')
            ws1[f'A{row}'].alignment = header_align
            ws1.row_dimensions[row].height = 25
            row += 1
            
            debt_data = [
                ("Umumiy qarz" if lang == "uz" else "Общий долг", total_debt, "so'm" if lang == "uz" else "сум"),
                ("Oddiy yo'l" if lang == "uz" else "Обычный путь", simple_months, "oy" if lang == "uz" else "мес"),
                ("HALOS bilan" if lang == "uz" else "С HALOS", pro_months, "oy" if lang == "uz" else "мес"),
                ("Tejagan vaqt" if lang == "uz" else "Сэкономлено", months_saved, "oy" if lang == "uz" else "мес"),
            ]
            
            for label, value, unit in debt_data:
                ws1[f'A{row}'] = label
                ws1[f'A{row}'].font = label_font
                ws1[f'A{row}'].border = thin_border
                
                ws1[f'B{row}'] = value
                ws1[f'B{row}'].font = money_font if "HALOS" in label else value_font
                ws1[f'B{row}'].number_format = '#,##0'
                ws1[f'B{row}'].border = thin_border
                ws1[f'B{row}'].alignment = Alignment(horizontal='right')
                
                ws1[f'C{row}'] = unit
                ws1[f'C{row}'].border = thin_border
                row += 1
        
        # ==================== SHEET 2: TRANSACTIONS ====================
        ws2 = wb.create_sheet("Tranzaksiyalar" if lang == "uz" else "Транзакции")
        
        ws2.column_dimensions['A'].width = 25
        ws2.column_dimensions['B'].width = 18
        ws2.column_dimensions['C'].width = 15
        
        # Header
        ws2.merge_cells('A1:C1')
        ws2['A1'] = "📋 OYLIK TRANZAKSIYALAR" if lang == "uz" else "📋 ТРАНЗАКЦИИ ЗА МЕСЯЦ"
        ws2['A1'].font = title_font
        ws2['A1'].alignment = title_align
        
        row = 3
        
        # Summary
        ws2[f'A{row}'] = "Jami daromad" if lang == "uz" else "Всего доход"
        ws2[f'B{row}'] = monthly_income_tx
        ws2[f'B{row}'].number_format = '#,##0'
        ws2[f'B{row}'].font = money_font
        row += 1
        
        ws2[f'A{row}'] = "Jami xarajat" if lang == "uz" else "Всего расход"
        ws2[f'B{row}'] = monthly_expense
        ws2[f'B{row}'].number_format = '#,##0'
        ws2[f'B{row}'].font = danger_font
        row += 1
        
        balance = monthly_income_tx - monthly_expense
        ws2[f'A{row}'] = "Balans" if lang == "uz" else "Баланс"
        ws2[f'B{row}'] = balance
        ws2[f'B{row}'].number_format = '#,##0'
        ws2[f'B{row}'].font = money_font if balance >= 0 else danger_font
        row += 2
        
        # Expenses by category
        if expense_by_cat:
            ws2[f'A{row}'] = "XARAJATLAR KATEGORIYA BO'YICHA" if lang == "uz" else "РАСХОДЫ ПО КАТЕГОРИЯМ"
            ws2[f'A{row}'].font = section_font
            ws2[f'A{row}'].fill = section_fill
            ws2.merge_cells(f'A{row}:C{row}')
            row += 1
            
            # Category names mapping
            from app.ai_assistant import EXPENSE_CATEGORIES
            cat_names = EXPENSE_CATEGORIES.get(lang, EXPENSE_CATEGORIES["uz"])
            
            sorted_cats = sorted(expense_by_cat.items(), key=lambda x: x[1], reverse=True)
            for cat, amount in sorted_cats:
                cat_name = cat_names.get(cat, cat)
                pct = (amount / monthly_expense * 100) if monthly_expense > 0 else 0
                
                ws2[f'A{row}'] = cat_name
                ws2[f'A{row}'].border = thin_border
                
                ws2[f'B{row}'] = amount
                ws2[f'B{row}'].number_format = '#,##0'
                ws2[f'B{row}'].border = thin_border
                ws2[f'B{row}'].alignment = Alignment(horizontal='right')
                
                ws2[f'C{row}'] = f"{pct:.1f}%"
                ws2[f'C{row}'].border = thin_border
                ws2[f'C{row}'].alignment = Alignment(horizontal='center')
                row += 1
        
        # ==================== SHEET 3: KUNLIK HISOBCHI ====================
        # User's requested daily expense tracker format
        ws3 = wb.create_sheet("Kunlik hisobchi" if lang == "uz" else "Дневник расходов")
        
        # Column widths
        ws3.column_dimensions['A'].width = 6
        ws3.column_dimensions['B'].width = 18
        ws3.column_dimensions['C'].width = 18
        ws3.column_dimensions['D'].width = 30
        
        # Yellow header style
        header_yellow = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        header_font_black = Font(name='Arial', size=11, bold=True)
        
        # Row 1: Headers
        headers = [
            ("№", 'A'),
            ("Kirim" if lang == "uz" else "Приход", 'B'),
            ("Chiqim" if lang == "uz" else "Расход", 'C'),
            ("Izoh" if lang == "uz" else "Комментарий", 'D')
        ]
        
        for header_text, col in headers:
            ws3[f'{col}1'] = header_text
            ws3[f'{col}1'].font = header_font_black
            ws3[f'{col}1'].fill = header_yellow
            ws3[f'{col}1'].border = thin_border
            ws3[f'{col}1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Get detailed transactions
        from app.ai_assistant import get_user_transactions, INCOME_CATEGORIES, EXPENSE_CATEGORIES
        transactions = await get_user_transactions(db, user["id"], days=30)
        
        # Row 2: Kun boshiga qoldiq (starting balance)
        ws3.merge_cells('A2:C2')
        ws3['A2'] = "Kun boshiga qoldiq" if lang == "uz" else "Остаток на начало"
        ws3['A2'].font = header_font_black
        ws3['A2'].border = thin_border
        ws3['A2'].alignment = Alignment(horizontal='center')
        
        # Calculate starting balance (profile based or estimated)
        starting_balance = free_cash  # Bo'sh pul = starting balance
        ws3['D2'] = starting_balance
        ws3['D2'].number_format = '#,##0.00'
        ws3['D2'].font = money_font
        ws3['D2'].fill = header_yellow
        ws3['D2'].border = thin_border
        ws3['D2'].alignment = Alignment(horizontal='right')
        
        # Data rows (35 rows like user's example)
        row = 3
        total_income_tracker = 0
        total_expense_tracker = 0
        
        # Sort transactions by date
        sorted_transactions = sorted(transactions, key=lambda x: x.get('created_at', ''), reverse=False)
        
        for i in range(1, 36):  # 35 rows
            ws3[f'A{row}'] = i
            ws3[f'A{row}'].border = thin_border
            ws3[f'A{row}'].alignment = Alignment(horizontal='center')
            
            ws3[f'B{row}'].border = thin_border
            ws3[f'B{row}'].number_format = '#,##0.00'
            ws3[f'B{row}'].alignment = Alignment(horizontal='right')
            
            ws3[f'C{row}'].border = thin_border
            ws3[f'C{row}'].number_format = '#,##0.00'
            ws3[f'C{row}'].alignment = Alignment(horizontal='right')
            
            ws3[f'D{row}'].border = thin_border
            
            # Fill with transaction data if available
            if i <= len(sorted_transactions):
                tx = sorted_transactions[i-1]
                tx_type = tx.get('type', '')
                tx_amount = tx.get('amount', 0)
                tx_category = tx.get('category', '')
                tx_desc = tx.get('description', '')
                
                # Get category name
                if tx_type == 'income':
                    cat_names = INCOME_CATEGORIES.get(lang, INCOME_CATEGORIES["uz"])
                    cat_display = cat_names.get(tx_category, tx_category)
                    ws3[f'B{row}'] = tx_amount
                    ws3[f'B{row}'].font = money_font
                    total_income_tracker += tx_amount
                else:
                    cat_names = EXPENSE_CATEGORIES.get(lang, EXPENSE_CATEGORIES["uz"])
                    cat_display = cat_names.get(tx_category, tx_category)
                    ws3[f'C{row}'] = tx_amount
                    ws3[f'C{row}'].font = danger_font
                    total_expense_tracker += tx_amount
                
                # Description: category + user description
                display_desc = f"{cat_display}"
                if tx_desc:
                    display_desc = tx_desc
                ws3[f'D{row}'] = display_desc
            
            row += 1
        
        # JAMI (Total) row
        ws3[f'A{row}'] = "JAMI" if lang == "uz" else "ИТОГО"
        ws3[f'A{row}'].font = header_font_black
        ws3[f'A{row}'].border = thin_border
        
        ws3[f'B{row}'] = total_income_tracker
        ws3[f'B{row}'].number_format = '#,##0.00'
        ws3[f'B{row}'].font = money_font
        ws3[f'B{row}'].border = thin_border
        ws3[f'B{row}'].alignment = Alignment(horizontal='right')
        
        ws3[f'C{row}'] = total_expense_tracker
        ws3[f'C{row}'].number_format = '#,##0.00'
        ws3[f'C{row}'].font = danger_font
        ws3[f'C{row}'].border = thin_border
        ws3[f'C{row}'].alignment = Alignment(horizontal='right')
        
        ws3[f'D{row}'].border = thin_border
        row += 1
        
        # Kun oxiriga qoldiq (ending balance)
        ws3.merge_cells(f'A{row}:C{row}')
        ws3[f'A{row}'] = "Kun oxiriga qoldiq" if lang == "uz" else "Остаток на конец"
        ws3[f'A{row}'].font = header_font_black
        ws3[f'A{row}'].border = thin_border
        ws3[f'A{row}'].alignment = Alignment(horizontal='center')
        
        ending_balance = starting_balance + total_income_tracker - total_expense_tracker
        ws3[f'D{row}'] = ending_balance
        ws3[f'D{row}'].number_format = '#,##0.00'
        ws3[f'D{row}'].font = money_font if ending_balance >= 0 else danger_font
        ws3[f'D{row}'].fill = header_yellow
        ws3[f'D{row}'].border = thin_border
        ws3[f'D{row}'].alignment = Alignment(horizontal='right')
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Send file using InputFile for proper BytesIO handling
        from telegram import InputFile
        filename = f"HALOS_Report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        
        if query:
            chat_id = query.message.chat_id
        else:
            chat_id = update.message.chat_id
        
        caption_uz = "📊 *HALOS Moliyaviy Hisobot*\n\n✅ 3 ta sahifa:\n• Xulosa\n• Tranzaksiyalar\n• Kunlik hisobchi"
        caption_ru = "📊 *HALOS Финансовый Отчёт*\n\n✅ 3 листа:\n• Сводка\n• Транзакции\n• Дневник расходов"
        
        await context.bot.send_document(
            chat_id=chat_id,
            document=InputFile(output, filename=filename),
            caption=caption_uz if lang == "uz" else caption_ru,
            parse_mode="Markdown"
        )
        
        # Update the loading message to success
        if lang == "uz":
            msg = "✅ *Excel hisobot yuborildi!*\n\n📋 Hisobotda:\n• Daromadlar va xarajatlar\n• HALOS usuli taqsimoti\n• Qarzdan chiqish muddati\n• Oylik tranzaksiyalar\n• Kunlik hisobchi"
        else:
            msg = "✅ *Excel отчёт отправлен!*\n\n📋 В отчёте:\n• Доходы и расходы\n• Распределение по методу HALOS\n• Срок выхода из долга\n• Транзакции за месяц\n• Дневник расходов"
        
        if query:
            try:
                await query.edit_message_text(msg, parse_mode="Markdown")
            except Exception:
                pass  # Message may have been superseded
        else:
            await update.message.reply_text(msg, parse_mode="Markdown")
            
    except Exception as e:
        logger.error(f"Excel export error: {e}")
        import traceback
        traceback.print_exc()
        if lang == "uz":
            msg = "❌ Eksportda xatolik yuz berdi. Keyinroq urinib ko'ring."
        else:
            msg = "❌ Ошибка при экспорте. Попробуйте позже."
        
        if query:
            await query.edit_message_text(msg)
        else:
            await update.message.reply_text(msg)


# ==================== PRO MENU ====================

async def show_pro_menu(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Show PRO features menu"""
    query = update.callback_query
    if query:
        await query.answer()
    
    telegram_id = update.effective_user.id
    lang = context.user_data.get("lang", "uz")
    
    # Check PRO status
    is_pro = await is_user_pro(telegram_id)
    
    if not is_pro:
        # Show pricing instead
        from app.subscription_handlers import show_pricing
        await show_pricing(update, context)
        return
    
    if lang == "uz":
        msg = (
            "💎 *HALOS PRO*\n"
            "━━━━━━━━━━━━━━━━━━━━━━\n\n"
            
            "✅ Siz PRO foydalanuvchisiz!\n\n"
            
            "🎯 *Sizning imkoniyatlaringiz:*\n\n"
            
            "🎤 *Ovozli AI yordamchi*\n"
            "   Ovoz yoki matn bilan xarajat yozing\n\n"
            
            "📊 *Batafsil statistika*\n"
            "   Haftalik, oylik, yillik tahlil\n\n"
            
            "📅 *Erkinlik sanangiz*\n"
            "   Qarzlardan qachon xalos bo'lasiz\n\n"
            
            "🔔 *Aqlli eslatmalar*\n"
            "   To'lov eslatmalari va maslahatlar\n\n"
            
            "📋 *Yuk nazorati*\n"
            "   Moliyaviy yuk monitoringi\n\n"
            
            "📥 *Excel hisobot*\n"
            "   Barcha ma'lumotlarni yuklab oling\n\n"
            
            "━━━━━━━━━━━━━━━━━━━━━━\n"
            "Quyidagi funksiyalardan birini tanlang:"
        )
    else:
        msg = (
            "💎 *HALOS PRO*\n"
            "━━━━━━━━━━━━━━━━━━━━━━\n\n"
            
            "✅ Вы PRO пользователь!\n\n"
            
            "🎯 *Ваши возможности:*\n\n"
            
            "🎤 *Голосовой AI помощник*\n"
            "   Записывайте расходы голосом или текстом\n\n"
            
            "📊 *Детальная статистика*\n"
            "   Еженедельный, ежемесячный, годовой анализ\n\n"
            
            "📅 *Дата Halos*\n"
            "   Когда освободитесь от долгов\n\n"
            
            "🔔 *Умные напоминания*\n"
            "   Напоминания о платежах и советы\n\n"
            
            "📋 *Контроль нагрузки*\n"
            "   Мониторинг финансовой нагрузки\n\n"
            
            "📥 *Excel отчёт*\n"
            "   Скачивайте все данные\n\n"
            
            "━━━━━━━━━━━━━━━━━━━━━━\n"
            "Выберите функцию:"
        )
    
    keyboard = [
        [InlineKeyboardButton(
            "🎤 Ovozli AI" if lang == "uz" else "🎤 Голосовой AI",
            callback_data="ai_assistant"
        )],
        [InlineKeyboardButton(
            "📊 Statistika" if lang == "uz" else "📊 Статистика",
            callback_data="pro_statistics"
        )],
        [InlineKeyboardButton(
            "🔔 Eslatmalar" if lang == "uz" else "🔔 Напоминания",
            callback_data="pro_reminders"
        )],
        [InlineKeyboardButton(
            "📋 Yuk nazorati" if lang == "uz" else "📋 Контроль нагрузки",
            callback_data="pro_debt_monitor"
        )],
        [InlineKeyboardButton(
            "📊 Hisobot sozlamalari" if lang == "uz" else "📊 Настройки отчётов",
            callback_data="report_settings"
        )],
        [InlineKeyboardButton(
            "📥 Excel yuklab olish" if lang == "uz" else "📥 Скачать Excel",
            callback_data="pro_export_excel"
        )],
        [InlineKeyboardButton(
            "◀️ Orqaga" if lang == "uz" else "◀️ Назад",
            callback_data="back_to_main"
        )]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    if query:
        await query.edit_message_text(msg, parse_mode="Markdown", reply_markup=reply_markup)
    else:
        await update.message.reply_text(msg, parse_mode="Markdown", reply_markup=reply_markup)


# ==================== CALLBACK HANDLERS ====================

async def pro_statistics_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handle statistics callback"""
    await show_statistics(update, context)


async def pro_reminders_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handle reminders callback"""
    await show_reminders(update, context)


async def pro_debt_monitor_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handle debt monitoring callback"""
    await show_debt_monitoring(update, context)


async def pro_export_excel_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handle Excel export callback"""
    await export_excel(update, context)


async def pro_menu_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handle PRO menu callback"""
    await show_pro_menu(update, context)


async def toggle_reminders_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handle reminder toggle"""
    query = update.callback_query
    await query.answer()
    
    lang = context.user_data.get("lang", "uz")
    action = query.data.replace("toggle_reminders_", "")
    
    if action == "on":
        if lang == "uz":
            msg = "✅ Eslatmalar yoqildi!"
        else:
            msg = "✅ Напоминания включены!"
    else:
        if lang == "uz":
            msg = "🔕 Eslatmalar o'chirildi."
        else:
            msg = "🔕 Напоминания отключены."
    
    await query.answer(msg, show_alert=True)
    
    # Go back to reminders menu
    await show_reminders(update, context)

# ==================== SCHEDULED REPORTS SETTINGS ====================

async def show_report_settings(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Show scheduled report settings"""
    query = update.callback_query
    if query:
        await query.answer()
    
    lang = context.user_data.get("lang", "uz")
    telegram_id = update.effective_user.id
    
    db = await get_database()
    user = await db.get_user(telegram_id)
    
    if not user:
        return
    
    # Get current settings
    daily = user.get("reports_daily", False)
    weekly = user.get("reports_weekly", False)
    monthly = user.get("reports_monthly", True)
    
    off_text = "O'chirilgan"
    off_btn = "O'chirish"
    on_btn = "Yoqish"
    
    if lang == "uz":
        msg = (
            "📊 *HISOBOT SOZLAMALARI*\n"
            "━━━━━━━━━━━━━━━━━━━━\n\n"
            "🔔 Avtomatik hisobotlar:\n\n"
            f"📅 *Kunlik hisobot:* {'✅ Yoqilgan' if daily else '❌ ' + off_text}\n"
            f"   _Har kuni soat 21:00 da_\n\n"
            f"📆 *Haftalik hisobot:* {'✅ Yoqilgan' if weekly else '❌ ' + off_text}\n"
            f"   _Har yakshanba soat 20:00 da_\n\n"
            f"🗓 *Oylik hisobot:* {'✅ Yoqilgan' if monthly else '❌ ' + off_text}\n"
            f"   _Har oyning 1-sanasi soat 19:00 da_\n\n"
            "━━━━━━━━━━━━━━━━━━━━\n"
            "💡 _Yoqish/o'chirish uchun tugmalarni bosing_"
        )
        
        keyboard = [
            [InlineKeyboardButton(
                f"{'🔕 ' + off_btn if daily else '✅ ' + on_btn} - Kunlik",
                callback_data=f"toggle_report_daily_{'off' if daily else 'on'}"
            )],
            [InlineKeyboardButton(
                f"{'🔕 ' + off_btn if weekly else '✅ ' + on_btn} - Haftalik",
                callback_data=f"toggle_report_weekly_{'off' if weekly else 'on'}"
            )],
            [InlineKeyboardButton(
                f"{'🔕 ' + off_btn if monthly else '✅ ' + on_btn} - Oylik",
                callback_data=f"toggle_report_monthly_{'off' if monthly else 'on'}"
            )],
            [InlineKeyboardButton("◀️ Orqaga", callback_data="pro_menu")]
        ]
    else:
        msg = (
            "📊 *НАСТРОЙКИ ОТЧЁТОВ*\n"
            "━━━━━━━━━━━━━━━━━━━━\n\n"
            "🔔 Автоматические отчёты:\n\n"
            f"📅 *Ежедневный:* {'✅ Включён' if daily else '❌ Выключен'}\n"
            f"   _Каждый день в 21:00_\n\n"
            f"📆 *Еженедельный:* {'✅ Включён' if weekly else '❌ Выключен'}\n"
            f"   _Каждое воскресенье в 20:00_\n\n"
            f"🗓 *Ежемесячный:* {'✅ Включён' if monthly else '❌ Выключен'}\n"
            f"   _1-го числа каждого месяца в 19:00_\n\n"
            "━━━━━━━━━━━━━━━━━━━━\n"
            "💡 _Нажмите кнопки для настройки_"
        )
        
        keyboard = [
            [InlineKeyboardButton(
                f"{'🔕 Выкл' if daily else '✅ Вкл'} - Ежедневный",
                callback_data=f"toggle_report_daily_{'off' if daily else 'on'}"
            )],
            [InlineKeyboardButton(
                f"{'🔕 Выкл' if weekly else '✅ Вкл'} - Еженедельный",
                callback_data=f"toggle_report_weekly_{'off' if weekly else 'on'}"
            )],
            [InlineKeyboardButton(
                f"{'🔕 Выкл' if monthly else '✅ Вкл'} - Ежемесячный",
                callback_data=f"toggle_report_monthly_{'off' if monthly else 'on'}"
            )],
            [InlineKeyboardButton("◀️ Назад", callback_data="pro_menu")]
        ]
    
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    if query:
        await query.edit_message_text(msg, parse_mode="Markdown", reply_markup=reply_markup)
    else:
        await update.message.reply_text(msg, parse_mode="Markdown", reply_markup=reply_markup)


async def toggle_report_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Toggle report setting on/off"""
    query = update.callback_query
    await query.answer()
    
    telegram_id = update.effective_user.id
    lang = context.user_data.get("lang", "uz")
    
    # Parse callback data: toggle_report_{period}_{on/off}
    data = query.data.replace("toggle_report_", "")
    parts = data.split("_")
    period = parts[0]  # daily, weekly, monthly
    action = parts[1]  # on, off
    
    new_value = action == "on"
    column = f"reports_{period}"
    
    db = await get_database()
    
    try:
        if db.is_postgres:
            async with db._pool.acquire() as conn:
                await conn.execute(f"""
                    UPDATE users SET {column} = $1 WHERE telegram_id = $2
                """, new_value, telegram_id)
        else:
            await db._connection.execute(f"""
                UPDATE users SET {column} = ? WHERE telegram_id = ?
            """, (1 if new_value else 0, telegram_id))
            await db._connection.commit()
        
        # Show success message
        if lang == "uz":
            period_names = {"daily": "Kunlik", "weekly": "Haftalik", "monthly": "Oylik"}
            if new_value:
                msg = f"✅ {period_names[period]} hisobot yoqildi!"
            else:
                msg = f"🔕 {period_names[period]} hisobot o'chirildi."
        else:
            period_names = {"daily": "Ежедневный", "weekly": "Еженедельный", "monthly": "Ежемесячный"}
            if new_value:
                msg = f"✅ {period_names[period]} отчёт включён!"
            else:
                msg = f"🔕 {period_names[period]} отчёт выключен."
        
        await query.answer(msg, show_alert=True)
        
    except Exception as e:
        logger.error(f"[REPORT] Toggle error: {e}")
        await query.answer("❌ Xatolik yuz berdi" if lang == "uz" else "❌ Произошла ошибка")
    
    # Refresh settings page
    await show_report_settings(update, context)


async def report_settings_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Callback for report settings button"""
    await show_report_settings(update, context)